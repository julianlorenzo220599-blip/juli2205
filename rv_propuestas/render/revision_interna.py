"""Renderizador Excel de revisión interna para Gabriel.

Hoja 1 "Resumen": KPIs + decisiones de sizing + cobertura.
Hoja 2 "BOM_Detalle": cada línea con costo, margen %, margen USD, precio con margen, IVA, final.
Hoja 3 "Resumen_Costos": totales por categoría con desglose márgenes / contingencia / financiero / IVA.
Hoja 4 "Notas": alertas del motor de sizing.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..config import CONTINGENCIA, COSTO_FINANCIERO, MARGENES, fmt_ar
from ..costeo.calculo import CosteoResult
from ..inputs.facturas import Factura
from ..inputs.pdi import PDI
from ..inputs.ubicacion import IrradiacionAnual
from ..sizing.engine import SizingResult
from ..sizing.topologia import ConfiguracionInversores, ConfiguracionStrings


THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
SUB_FONT = Font(bold=True, color="1F3864")
TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
LABEL_FONT = Font(bold=True, size=11, color="1F4E78")
VALUE_FONT = Font(size=11)
ALERT_FILL = PatternFill("solid", fgColor="FFE699")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def _hdr(ws, row: int, headers: list[str]):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def _autosize(ws, widths: list[int]):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def render(
    factura: Factura,
    irradiacion: IrradiacionAnual,
    pdi: PDI,
    sizing: SizingResult,
    inv_cfg: ConfiguracionInversores,
    str_cfg: ConfiguracionStrings,
    costeo: CosteoResult,
    output_path: str | Path,
    cliente_nombre: str = "",
    proyecto_nombre: str = "",
) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    ws["B2"] = f"REVISIÓN INTERNA — {proyecto_nombre or 'Proyecto'}"
    ws["B2"].font = TITLE_FONT
    ws.merge_cells("B2:F2")

    ws["B3"] = f"Cliente: {cliente_nombre or factura.titular or '—'}"
    ws["B3"].font = Font(italic=True, size=10)
    ws.merge_cells("B3:F3")

    row = 5

    def kv(label: str, value: str):
        nonlocal row
        ws.cell(row=row, column=2, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2).alignment = RIGHT
        ws.cell(row=row, column=3, value=value).font = VALUE_FONT
        ws.cell(row=row, column=3).alignment = LEFT
        row += 1

    ws.cell(row=row, column=2, value="DATOS DEL CLIENTE").fill = SUB_FILL
    ws.cell(row=row, column=2).font = SUB_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    row += 1
    kv("Distribuidora:", factura.distribuidora)
    kv("Categoría tarifaria:", factura.categoria_tarifaria)
    kv("Tensión suministro:", factura.tension_suministro)
    kv("Pot. contratada:", f"{factura.potencia_contratada_kw or '—'} kW")
    kv("Consumo anual:", f"{fmt_ar(factura.consumo_anual_kwh, 0)} kWh")
    kv("NIS:", factura.nis or "—")

    row += 1
    ws.cell(row=row, column=2, value="PDI").fill = SUB_FILL
    ws.cell(row=row, column=2).font = SUB_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    row += 1
    kv("Configuración:", pdi.descripcion)
    kv("Capacidad disponible:", f"{pdi.capacidad_disponible_kw or '—'} kW")
    kv("Distancia tablero:", f"{pdi.distancia_al_tablero_m} m")
    kv("Requiere trafo:", "Sí" if pdi.requiere_trafo_elevador else "No")

    row += 1
    ws.cell(row=row, column=2, value="SIZING").fill = SUB_FILL
    ws.cell(row=row, column=2).font = SUB_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    row += 1
    kv("Ubicación:", f"{irradiacion.ubicacion.nombre} ({irradiacion.ubicacion.lat:.3f}, {irradiacion.ubicacion.lon:.3f})")
    kv("Gen. específica:", f"{fmt_ar(irradiacion.generacion_especifica_kwh_kwp, 0)} kWh/kWp·año ({irradiacion.fuente})")
    kv("kWp recomendado:", f"{fmt_ar(sizing.kwp_real, 2)} kWp")
    kv("N° paneles 725W:", f"{sizing.n_paneles}")
    kv("Inversores:", f"{inv_cfg.cantidad} × {inv_cfg.inversor.descripcion}")
    kv("DC/AC:", f"{inv_cfg.ratio_dc_ac}")
    kv("Strings:", f"{str_cfg.n_strings} strings × {str_cfg.n_paneles_por_string} paneles")
    kv("Voc string (frío):", f"{str_cfg.voc_string} V (ventana inv: {inv_cfg.inversor.v_mppt_min}-{inv_cfg.inversor.v_mppt_max} V)")
    kv("Generación anual:", f"{fmt_ar(sizing.generacion_anual_kwh, 0)} kWh")
    kv("Cobertura:", f"{sizing.cobertura * 100:.1f}% (objetivo {sizing.cobertura_objetivo * 100:.0f}%)")
    kv("Topología:", f"{sizing.topologia} — {sizing.topologia_descripcion}")
    if sizing.limitado_por_pdi:
        c = ws.cell(row=row, column=2, value="⚠ Limitado por PDI:")
        c.font = LABEL_FONT
        c.fill = ALERT_FILL
        ws.cell(row=row, column=3, value=f"Objetivo sin tope: {sizing.kwp_solicitado_sin_limite} kWp").alignment = LEFT
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="KPI ECONÓMICOS (USD)").fill = SUB_FILL
    ws.cell(row=row, column=2).font = SUB_FONT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
    row += 1
    kv("Costo neto BOM:", f"USD {fmt_ar(costeo.costo_neto_total, 2)}")
    kv("Subtotal con margen:", f"USD {fmt_ar(costeo.subtotal_con_margen, 2)}")
    kv(f"Contingencia ({CONTINGENCIA*100:.0f}%):", f"USD {fmt_ar(costeo.contingencia_usd, 2)}")
    kv(f"Costo financiero ({COSTO_FINANCIERO*100:.0f}%):", f"USD {fmt_ar(costeo.costo_financiero_usd, 2)}")
    kv("Neto cliente (sin IVA):", f"USD {fmt_ar(costeo.neto_cliente, 2)}")
    kv("IVA total:", f"USD {fmt_ar(costeo.iva_total, 2)}")
    kv("TOTAL CLIENTE (USD):", f"USD {fmt_ar(costeo.total_cliente, 2)}")
    ws.cell(row=row - 1, column=3).font = Font(bold=True, size=12, color="C00000")
    kv("USD/kWp:", f"USD {fmt_ar(costeo.total_cliente / sizing.kwp_real, 0)}")

    _autosize(ws, [2, 28, 60, 4])

    # ── Hoja 2: BOM_Detalle ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("BOM_Detalle")
    hdr = ["Categoría", "SKU", "Descripción", "Cant.", "Unidad", "Precio unit. USD",
           "Costo neto USD", "Margen %", "Margen USD", "Con margen USD", "IVA %", "IVA USD", "Precio final USD", "Notas"]
    _hdr(ws2, 1, hdr)
    r = 2
    for linea in costeo.lineas:
        it = linea.item
        ws2.cell(row=r, column=1, value=it.categoria)
        ws2.cell(row=r, column=2, value=it.sku)
        ws2.cell(row=r, column=3, value=it.descripcion)
        ws2.cell(row=r, column=4, value=it.cantidad)
        ws2.cell(row=r, column=5, value=it.unidad)
        ws2.cell(row=r, column=6, value=it.precio_unit_usd if it.precio_unit_usd is not None else "—")
        ws2.cell(row=r, column=7, value=linea.costo_neto)
        ws2.cell(row=r, column=8, value=linea.margen_pct)
        ws2.cell(row=r, column=8).number_format = "0%"
        ws2.cell(row=r, column=9, value=linea.margen_usd)
        ws2.cell(row=r, column=10, value=linea.precio_con_margen)
        ws2.cell(row=r, column=11, value=linea.iva_pct)
        ws2.cell(row=r, column=11).number_format = "0.0%"
        ws2.cell(row=r, column=12, value=linea.iva_usd)
        ws2.cell(row=r, column=13, value=linea.precio_final)
        ws2.cell(row=r, column=14, value=it.notas or "")
        for c in range(1, len(hdr) + 1):
            ws2.cell(row=r, column=c).border = BORDER
            if c in (3, 14):
                ws2.cell(row=r, column=c).alignment = LEFT
            else:
                ws2.cell(row=r, column=c).alignment = CENTER
        if it.precio_unit_usd is None:
            ws2.cell(row=r, column=6).fill = ALERT_FILL
        r += 1
    _autosize(ws2, [12, 16, 50, 8, 8, 14, 14, 9, 14, 16, 7, 12, 16, 24])
    ws2.freeze_panes = "A2"

    # ── Hoja 3: Resumen_Costos ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Resumen_Costos")
    _hdr(ws3, 1, ["Categoría", "Margen aplicado", "Costo neto USD", "Margen USD", "Precio con margen USD"])
    r = 2
    for tc in costeo.totales_categoria:
        ws3.cell(row=r, column=1, value=tc.categoria)
        ws3.cell(row=r, column=2, value=MARGENES.get(tc.categoria, 0))
        ws3.cell(row=r, column=2).number_format = "0%"
        ws3.cell(row=r, column=3, value=tc.costo_neto)
        ws3.cell(row=r, column=4, value=tc.margen_usd)
        ws3.cell(row=r, column=5, value=tc.precio_con_margen)
        for c in range(1, 6):
            ws3.cell(row=r, column=c).border = BORDER
            ws3.cell(row=r, column=c).alignment = CENTER
        r += 1
    # Totales
    r += 1
    ws3.cell(row=r, column=1, value="Subtotal con margen").font = LABEL_FONT
    ws3.cell(row=r, column=5, value=costeo.subtotal_con_margen).font = Font(bold=True)
    r += 1
    ws3.cell(row=r, column=1, value=f"+ Contingencia ({CONTINGENCIA*100:.0f}%)").font = LABEL_FONT
    ws3.cell(row=r, column=5, value=costeo.contingencia_usd)
    r += 1
    ws3.cell(row=r, column=1, value=f"+ Costo financiero ({COSTO_FINANCIERO*100:.0f}%)").font = LABEL_FONT
    ws3.cell(row=r, column=5, value=costeo.costo_financiero_usd)
    r += 1
    ws3.cell(row=r, column=1, value="= Neto cliente").font = LABEL_FONT
    ws3.cell(row=r, column=5, value=costeo.neto_cliente).font = Font(bold=True)
    r += 1
    ws3.cell(row=r, column=1, value="+ IVA").font = LABEL_FONT
    ws3.cell(row=r, column=5, value=costeo.iva_total)
    r += 1
    ws3.cell(row=r, column=1, value="TOTAL CLIENTE").font = Font(bold=True, color="C00000")
    ws3.cell(row=r, column=5, value=costeo.total_cliente).font = Font(bold=True, color="C00000", size=12)
    _autosize(ws3, [30, 16, 18, 18, 24])

    # ── Hoja 4: Notas ─────────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Notas")
    ws4["A1"] = "ALERTAS Y NOTAS DEL MOTOR"
    ws4["A1"].font = TITLE_FONT
    r = 3
    if costeo.items_sin_precio:
        ws4.cell(row=r, column=1, value="⚠ ITEMS SIN PRECIO EN CATÁLOGO:").font = Font(bold=True, color="C00000")
        r += 1
        for sku in costeo.items_sin_precio:
            ws4.cell(row=r, column=1, value=f"  • {sku}")
            r += 1
        r += 1
    if sizing.notas:
        ws4.cell(row=r, column=1, value="NOTAS DE SIZING:").font = Font(bold=True)
        r += 1
        for n in sizing.notas:
            ws4.cell(row=r, column=1, value=f"  • {n}")
            ws4.cell(row=r, column=1).alignment = LEFT
            r += 1
    if not str_cfg.dentro_ventana_mppt:
        ws4.cell(row=r, column=1, value="⚠ STRING SIZING FUERA DE VENTANA MPPT — revisar paneles/string").fill = ALERT_FILL
        r += 1
    ws4.column_dimensions["A"].width = 110

    wb.save(output_path)
    return output_path
