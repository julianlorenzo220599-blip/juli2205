"""Build reorganized Despiece workbook with dashboard (SOLARMET + MIBET)."""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SRC = "./DESPIECE LISTA COMPLETA 3.xlsx"
DST = "./output/DESPIECE_KITS_DASHBOARD.xlsx"

KITS = [
    ("EC-3K",   "Economy 3 kW",                  "Monofásica", "On-Grid", 3,  6,  "N"),
    ("EC-5K",   "Economy Plus 5 kW",             "Monofásica", "On-Grid", 5,  10, "N"),
    ("EC-6K",   "Economy Pro 6 kW",              "Monofásica", "On-Grid", 6,  12, "N"),
    ("AC/R-3K", "Always Connected / Rural 3 kW", "Monofásica", "Híbrido", 3,  6,  "N"),
    ("AC/R-5K", "Always Connected / Rural 5 kW", "Monofásica", "Híbrido", 5,  12, "N"),
    ("ECT-8K",  "Economy T 8 kW",                "Trifásica",  "On-Grid", 8,  16, "N"),
    ("ECT-12K", "Economy T Plus 12 kW",          "Trifásica",  "On-Grid", 12, 24, "N"),
    ("ECT-20K", "Economy T Pro 20 kW",           "Trifásica",  "On-Grid", 20, 40, "N"),
    ("PS-8K",   "Power Station 8 kW",            "Trifásica",  "Híbrido", 8,  16, "S"),
    ("PS-10K",  "Power Station Plus 10 kW",      "Trifásica",  "Híbrido", 10, 18, "S"),
    ("PS-12K",  "Power Station Pro 12 kW",       "Trifásica",  "Híbrido", 12, 24, "S"),
    ("PS-15K",  "Power Station Mega 15 kW",      "Trifásica",  "Híbrido", 15, 32, "S"),
]

INV = {
    "EC-3K":   (1, "Inversor Monofásico ON-GRID 3 kW", "Inverters", "GOODWE", "GW3000-XS-30"),
    "EC-5K":   (1, "Inversor Monofásico ON-GRID 5 kW", "Inverters", "GOODWE", "GW5000-MS-30"),
    "EC-6K":   (1, "Inversor Monofásico ON-GRID 6 kW", "Inverters", "GOODWE", "GW6000-MS-30"),
    "AC/R-3K": (1, "Inversor Monofásico HIBRIDO 3 kW", "Inverters", "GOODWE", "GW3000-ES-20"),
    "AC/R-5K": (1, "Inversor Monofásico HIBRIDO 5 kW", "Inverters", "GOODWE", "GW5000-ES-20"),
    "ECT-8K":  (1, "Inversor Trifásico ON-GRID 8 kW",  "Inverters", "GOODWE", "GW8000-SDT-30"),
    "ECT-12K": (1, "Inversor Trifásico ON-GRID 12 kW", "Inverters", "GOODWE", "GW12K-SDT-30"),
    "ECT-20K": (1, "Inversor Trifásico ON-GRID 20 kW", "Inverters", "GOODWE", "GW20K-SDT-30"),
    "PS-8K":   (1, "Inversor Trifásico HIBRIDO 8 kW",  "Inverters", "GOODWE", "GW8000-ET-20"),
    "PS-10K":  (1, "Inversor Trifásico HIBRIDO 10 kW", "Inverters", "GOODWE", "GW10K-ET-20"),
    "PS-12K":  (1, "Inversor Trifásico HIBRIDO 12 kW", "Inverters", "GOODWE", "GW12K-ET-20"),
    "PS-15K":  (1, "Inversor Trifásico HIBRIDO 15 kW", "Inverters", "GOODWE", "GW15K-ET-20"),
}
PANEL  = ("Módulo Fotovoltaico TOPCon Bifacial 585w", "Paneles FV", "TCL SOLAR", "TCL-MI585DH182-72NT")
CABLE  = ("SET Cables + Conectores", "Cables", "GENERICO", "CC-20")
GMK110 = (1, "Smart Energy Controller Monofásico", "Accesorios", "GOODWE", "GMK110")
GMK330 = (1, "Smart Energy Controller Trifásico",  "Accesorios", "GOODWE", "GMK330")
BAT_LV = ("Batería 5kWh - 48V (con supresión de fuego)", "Baterías", "GOODWE", "LX U5.0-30")
BAT_HV = ("Batería Apilable 5kWh - HV", "Baterías", "GOODWE", "LX D5.0-10")

BAT = {
    "EC-3K":   (1, 0, None, GMK110),
    "EC-5K":   (2, 0, None, GMK110),
    "EC-6K":   (2, 0, None, GMK110),
    "AC/R-3K": (1, 1, "LV", GMK110),
    "AC/R-5K": (2, 1, "LV", GMK110),
    "ECT-8K":  (2, 0, None, GMK330),
    "ECT-12K": (3, 0, None, GMK330),
    "ECT-20K": (5, 0, None, GMK330),
    "PS-8K":   (2, 1, "HV", GMK330),
    "PS-10K":  (3, 1, "HV", GMK330),
    "PS-12K":  (3, 2, "HV", GMK330),
    "PS-15K":  (4, 2, "HV", GMK330),
}

ESTR = {
    "EC-3K":   (1, 6), "EC-5K":   (2, 5), "EC-6K":   (2, 6),
    "AC/R-3K": (1, 6), "AC/R-5K": (2, 6),
    "ECT-8K":  (2, 8), "ECT-12K": (4, 6), "ECT-20K": (5, 8),
    "PS-8K":   (2, 8), "PS-10K":  (3, 6), "PS-12K":  (4, 6), "PS-15K":  (4, 8),
}

SOPORTE = {"PS-8K": 1, "PS-10K": 1, "PS-12K": 2, "PS-15K": 2}

# ---------- SOLARMET despiece (qty por set, item, código) ----------
SOLARMET_DETALLE = {
    "C-5U": [(12, "MiniMET 0.5m Mini Riel Solarmet", "NESTMNR050"),
             (1,  "Kit Anclajes Laterales",           "NESTSKITAL"),
             (2,  "Kit Anclajes Medios - 4 Paneles",  "NESTSKITM4")],
    "C-6U": [(14, "MiniMET 0.5m Mini Riel Solarmet", "NESTMNR050"),
             (1,  "Kit Anclajes Laterales",           "NESTSKITAL"),
             (2,  "Kit Anclajes Medios - 4 Paneles",  "NESTSKITM4")],
    "C-8U": [(18, "MiniMET 0.5m Mini Riel Solarmet", "NESTMNR050"),
             (1,  "Kit Anclajes Laterales",           "NESTSKITAL"),
             (3,  "Kit Anclajes Medios - 4 Paneles",  "NESTSKITM4")],
    "T-5U": [(4, "Perfil CODA N2 (3.55m)",          "NESTSCO35D"),
             (2, "Anclaje Teja V2",                 "NESTSTEJA1"),
             (1, "Kit Anclajes Laterales",          "NESTSKITAL"),
             (2, "Kit Anclajes Medios - 4 Paneles", "NESTSKITM4"),
             (1, "Kit Conexion CODA N2",            "NESTSKITCX")],
    "T-6U": [(4, "Perfil CODA N2 (3.55m)",          "NESTSCO35D"),
             (2, "Anclaje Teja V2",                 "NESTSTEJA1"),
             (1, "Kit Anclajes Laterales",          "NESTSKITAL"),
             (2, "Kit Anclajes Medios - 4 Paneles", "NESTSKITM4"),
             (1, "Kit Conexion CODA N2",            "NESTSKITCX")],
    "T-8U": [(6, "Perfil CODA N2 (3.55m)",          "NESTSCO35D"),
             (2, "Anclaje Teja V2",                 "NESTSTEJA1"),
             (1, "Kit Anclajes Laterales",          "NESTSKITAL"),
             (3, "Kit Anclajes Medios - 4 Paneles", "NESTSKITM4"),
             (2, "Kit Conexion CODA N2",            "NESTSKITCX")],
    "SL-5U": [(3, "Perfil CODA N2 (4.75m) V2",       "IESTSCO47D"),
              (1, "Kit Anclajes Laterales",          "NESTSKITAL"),
              (2, "Kit Anclajes Medios - 4 Paneles", "NESTSKITM4"),
              (1, "Kit Conexion CODA N2",            "NESTSKITCX"),
              (5, "Triangulo Aluminio Regulable",    "NESTSTRIAN")],
    "SL-6U": [(3, "Perfil CODA N2 (4.75m) V2",       "IESTSCO47D"),
              (1, "Kit Anclajes Laterales",          "NESTSKITAL"),
              (2, "Kit Anclajes Medios - 4 Paneles", "NESTSKITM4"),
              (1, "Kit Conexion CODA N2",            "NESTSKITCX"),
              (6, "Triangulo Aluminio Regulable",    "NESTSTRIAN")],
    "SL-8U": [(4, "Perfil CODA N2 (4.75m) V2",       "IESTSCO47D"),
              (1, "Kit Anclajes Laterales",          "NESTSKITAL"),
              (3, "Kit Anclajes Medios - 4 Paneles", "NESTSKITM4"),
              (1, "Kit Conexion CODA N2",            "NESTSKITCX"),
              (7, "Triangulo Aluminio Regulable",    "NESTSTRIAN")],
}

# ---------- MIBET despiece (qty por set, código, descripción) ----------
# Chapa y Teja. Suelo/Losa NO va acá: viene como kit completo GMS-3x720.
MIBET_DETALLE = {
    "S5-CH": [(2,  "AL6005-T5-22", "Riel aluminio 4.8m"),
              (2,  "AL6005-T5-24", "Riel aluminio 2.4m"),
              (2,  "AL6005-T5-23", "Empalme/unión de rieles"),
              (8,  "AL6005-T5-19", "Prensa intermedia (marco 30/35mm)"),
              (4,  "AL6005-T5-16", "Prensa final (marco 30/35mm)"),
              (10, "AL6005-T5-20", "Pata L para chapa trapezoidal/sinusoidal"),
              (5,  "SUS304-4",     "Clip de puesta a tierra (panel-riel)"),
              (1,  "AL6005-T5-10", "Terminal de tierra (riel-cable PE)"),
              (5,  "SUS304-6 (CC)","Clip para cable (gestión de cableado)")],
    "S6-CH": [(4,  "AL6005-T5-22", "Riel aluminio 4.8m"),
              (2,  "AL6005-T5-23", "Empalme/unión de rieles"),
              (10, "AL6005-T5-19", "Prensa intermedia (marco 30/35mm)"),
              (4,  "AL6005-T5-16", "Prensa final (marco 30/35mm)"),
              (12, "AL6005-T5-20", "Pata L para chapa trapezoidal/sinusoidal"),
              (6,  "SUS304-4",     "Clip de puesta a tierra (panel-riel)"),
              (1,  "AL6005-T5-10", "Terminal de tierra (riel-cable PE)"),
              (6,  "SUS304-6 (CC)","Clip para cable (gestión de cableado)")],
    "S8-CH": [(4,  "AL6005-T5-22", "Riel aluminio 4.8m"),
              (2,  "AL6005-T5-24", "Riel aluminio 2.4m"),
              (4,  "AL6005-T5-23", "Empalme/unión de rieles"),
              (14, "AL6005-T5-19", "Prensa intermedia (marco 30/35mm)"),
              (4,  "AL6005-T5-16", "Prensa final (marco 30/35mm)"),
              (16, "AL6005-T5-20", "Pata L para chapa trapezoidal/sinusoidal"),
              (8,  "SUS304-4",     "Clip de puesta a tierra (panel-riel)"),
              (1,  "AL6005-T5-10", "Terminal de tierra (riel-cable PE)"),
              (8,  "SUS304-6 (CC)","Clip para cable (gestión de cableado)")],
    "S5-TJ": [(2,  "AL6005-T5-22", "Riel aluminio 4.8m"),
              (2,  "AL6005-T5-24", "Riel aluminio 2.4m"),
              (2,  "AL6005-T5-23", "Empalme/unión de rieles"),
              (8,  "AL6005-T5-19", "Prensa intermedia (marco 30/35mm)"),
              (4,  "AL6005-T5-16", "Prensa final (marco 30/35mm)"),
              (10, "SUS304-7",     "Gancho para teja"),
              (5,  "SUS304-4",     "Clip de puesta a tierra (panel-riel)"),
              (1,  "AL6005-T5-10", "Terminal de tierra (riel-cable PE)"),
              (5,  "SUS304-6 (CC)","Clip para cable (gestión de cableado)")],
    "S6-TJ": [(4,  "AL6005-T5-22", "Riel aluminio 4.8m"),
              (2,  "AL6005-T5-23", "Empalme/unión de rieles"),
              (10, "AL6005-T5-19", "Prensa intermedia (marco 30/35mm)"),
              (4,  "AL6005-T5-16", "Prensa final (marco 30/35mm)"),
              (12, "SUS304-7",     "Gancho para teja"),
              (6,  "SUS304-4",     "Clip de puesta a tierra (panel-riel)"),
              (1,  "AL6005-T5-10", "Terminal de tierra (riel-cable PE)"),
              (6,  "SUS304-6 (CC)","Clip para cable (gestión de cableado)")],
    "S8-TJ": [(4,  "AL6005-T5-22", "Riel aluminio 4.8m"),
              (2,  "AL6005-T5-24", "Riel aluminio 2.4m"),
              (4,  "AL6005-T5-23", "Empalme/unión de rieles"),
              (14, "AL6005-T5-19", "Prensa intermedia (marco 30/35mm)"),
              (4,  "AL6005-T5-16", "Prensa final (marco 30/35mm)"),
              (16, "SUS304-7",     "Gancho para teja"),
              (8,  "SUS304-4",     "Clip de puesta a tierra (panel-riel)"),
              (1,  "AL6005-T5-10", "Terminal de tierra (riel-cable PE)"),
              (8,  "SUS304-6 (CC)","Clip para cable (gestión de cableado)")],
}

# Kits GMS-3x720 por set Suelo MIBET
MIBET_SUELO_KITS_POR_SET = {5: 1, 6: 1, 8: 2}

# Layout
BASE_ROWS_MAX = 5
DET_ROWS_MAX  = 9

# Styles
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
SUB_FONT = Font(bold=True, color="1F3864")
TITLE_FONT = Font(bold=True, size=16, color="1F4E78")
LABEL_FONT = Font(bold=True, size=11, color="1F4E78")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
INPUT_FONT = Font(bold=True, size=12)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def style_body(ws, start_row, end_row, ncols, item_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = LEFT if c == item_col else CENTER


os.makedirs(os.path.dirname(DST), exist_ok=True)
wb = openpyxl.Workbook()

# ---------- Kits ----------
ws_k = wb.active
ws_k.title = "Kits"
headers = ["Código Kit", "Nombre", "Fase", "Tipo", "Potencia (kW)", "Cant. Paneles", "Lleva Soporte Batería"]
ws_k.append(headers)
style_header(ws_k, 1, range(1, len(headers) + 1))
for k in KITS:
    ws_k.append(list(k))
style_body(ws_k, 2, len(KITS) + 1, len(headers), 2)
autosize(ws_k, [12, 32, 14, 12, 14, 14, 22])
ws_k.freeze_panes = "A2"
kits_last_row = len(KITS) + 1

# ---------- BOM_Base ----------
ws_b = wb.create_sheet("BOM_Base")
hdr = ["Key", "Código Kit", "Tipo", "Qty", "Item", "Categoría", "Marca", "Código"]
ws_b.append(hdr)
style_header(ws_b, 1, range(1, len(hdr) + 1))
for codigo, _, _, _, _, paneles, _ in KITS:
    inv_qty, inv_item, inv_cat, inv_mk, inv_cod = INV[codigo]
    qcab, qbat, tbat, smart = BAT[codigo]
    items = [
        (inv_qty, inv_item, inv_cat, inv_mk, inv_cod),
        (paneles, *PANEL),
        (qcab, *CABLE),
    ]
    if qbat > 0:
        bat = BAT_LV if tbat == "LV" else BAT_HV
        items.append((qbat, *bat))
    s_q, s_it, s_cat, s_mk, s_cod = smart
    items.append((s_q, s_it, s_cat, s_mk, s_cod))
    for idx, it in enumerate(items, 1):
        ws_b.append([f"{codigo}|{idx}", codigo, "Base", *it])
bom_base_last = ws_b.max_row
style_body(ws_b, 2, bom_base_last, len(hdr), 5)
autosize(ws_b, [18, 12, 10, 8, 42, 14, 12, 22])
ws_b.freeze_panes = "A2"

# ---------- BOM_Estructura (SOLARMET + MIBET) ----------
# Key = Kit|TipoEstructura|Marca
ws_e = wb.create_sheet("BOM_Estructura")
hdr_e = ["Key", "Código Kit", "Tipo Estructura", "Marca", "Qty", "Item", "Categoría", "Código Set"]
ws_e.append(hdr_e)
style_header(ws_e, 1, range(1, len(hdr_e) + 1))
TIPOS = [("Chapa", "SOLARMET_PFX_C", "MIBET_PFX_CH"),
         ("Teja", "SOLARMET_PFX_T", "MIBET_PFX_TJ"),
         ("Suelo/Losa", "SOLARMET_PFX_SL", "MIBET_PFX_SL")]
for k in KITS:
    codigo = k[0]
    qty_sets, size = ESTR[codigo]
    # Chapa
    sm_set = f"C-{size}U"; mb_set = f"S{size}-CH"
    ws_e.append([f"{codigo}|Chapa|SOLARMET", codigo, "Chapa", "SOLARMET", qty_sets,
                 f"Set Estructuras Techo Coplanar Chapa SOLARMET (x{size})", "Estructuras", sm_set])
    ws_e.append([f"{codigo}|Chapa|MIBET", codigo, "Chapa", "MIBET", qty_sets,
                 f"Set Estructura Coplanar Chapa MIBET (x{size})", "Estructuras", mb_set])
    # Teja
    sm_set = f"T-{size}U"; mb_set = f"S{size}-TJ"
    ws_e.append([f"{codigo}|Teja|SOLARMET", codigo, "Teja", "SOLARMET", qty_sets,
                 f"Set Estructuras Techo Coplanar Teja SOLARMET (x{size})", "Estructuras", sm_set])
    ws_e.append([f"{codigo}|Teja|MIBET", codigo, "Teja", "MIBET", qty_sets,
                 f"Set Estructura Coplanar Teja MIBET (x{size})", "Estructuras", mb_set])
    # Suelo/Losa
    sm_set = f"SL-{size}U"
    ws_e.append([f"{codigo}|Suelo/Losa|SOLARMET", codigo, "Suelo/Losa", "SOLARMET", qty_sets,
                 f"Set Estructuras Suelo / Losa SOLARMET (x{size})", "Estructuras", sm_set])
    # MIBET Suelo: kit GMS-3x720 indivisible. Qty = sets × kits-per-set
    kits_por_set = MIBET_SUELO_KITS_POR_SET[size]
    qty_kits = qty_sets * kits_por_set
    ws_e.append([f"{codigo}|Suelo/Losa|MIBET", codigo, "Suelo/Losa", "MIBET", qty_kits,
                 f"Kit Estructura Suelo MIBET GMS-3x720 (6 paneles c/u — viene armado)",
                 "Estructuras", "GMS-3x720"])
bom_est_last = ws_e.max_row
style_body(ws_e, 2, bom_est_last, len(hdr_e), 6)
autosize(ws_e, [28, 12, 14, 10, 8, 48, 14, 14])
ws_e.freeze_panes = "A2"

# ---------- BOM_Soporte ----------
ws_s = wb.create_sheet("BOM_Soporte")
hdr_s = ["Key", "Código Kit", "Tipo Soporte", "Qty", "Item", "Categoría", "Marca", "Código"]
ws_s.append(hdr_s)
style_header(ws_s, 1, range(1, len(hdr_s) + 1))
for codigo, qty in SOPORTE.items():
    ws_s.append([f"{codigo}|Pared", codigo, "Pared", qty,
                 "Soporte batería Lynx D para pared", "Accesorios", "SOLARMET", "Hanger assembly for Lynx D"])
    ws_s.append([f"{codigo}|Suelo", codigo, "Suelo", qty,
                 "Base batería Lynx D para suelo", "Accesorios", "SOLARMET", "Base assembly for Lynx D"])
bom_sop_last = ws_s.max_row
style_body(ws_s, 2, bom_sop_last, len(hdr_s), 5)
autosize(ws_s, [18, 12, 14, 8, 42, 14, 12, 28])
ws_s.freeze_panes = "A2"

# ---------- Estructuras_Detalle (SOLARMET + MIBET Chapa/Teja). MIBET Suelo no aplica. ----------
ws_ed = wb.create_sheet("Estructuras_Detalle")
hdr_ed = ["Key", "Código Set", "Marca", "Qty (por set)", "Item", "Categoría", "Código"]
ws_ed.append(hdr_ed)
style_header(ws_ed, 1, range(1, len(hdr_ed) + 1))
for set_code, comps in SOLARMET_DETALLE.items():
    for idx, (q, it, cod) in enumerate(comps, 1):
        ws_ed.append([f"{set_code}|{idx}", set_code, "SOLARMET", q, it, "Estructuras", cod])
for set_code, comps in MIBET_DETALLE.items():
    for idx, (q, cod, it) in enumerate(comps, 1):
        ws_ed.append([f"{set_code}|{idx}", set_code, "MIBET", q, it, "Estructuras", cod])
ed_last = ws_ed.max_row
style_body(ws_ed, 2, ed_last, len(hdr_ed), 5)
autosize(ws_ed, [12, 12, 10, 14, 42, 14, 16])
ws_ed.freeze_panes = "A2"

# ---------- Master_Data ----------
ws_md = wb.create_sheet("Master_Data")
loaded_md = False
if os.path.exists(SRC):
    try:
        src = openpyxl.load_workbook(SRC, data_only=True)
        if "Listado Master DATA" in src.sheetnames:
            src_md = src["Listado Master DATA"]
            for row in src_md.iter_rows(values_only=True):
                ws_md.append(row)
            style_header(ws_md, 1, range(1, src_md.max_column + 1))
            style_body(ws_md, 2, ws_md.max_row, src_md.max_column, 4)
            autosize(ws_md, [14, 14, 28, 60] + [14] * max(0, src_md.max_column - 4))
            loaded_md = True
    except Exception as exc:
        print("Master_Data load failed:", exc)
if not loaded_md:
    ws_md.append(["Categoría", "Marca", "Código (SKU)", "Descripción"])
    style_header(ws_md, 1, range(1, 5))
    autosize(ws_md, [14, 14, 28, 60])
ws_md.freeze_panes = "A2"

# ---------- Dashboard ----------
ws_d = wb.create_sheet("Dashboard", 0)

ws_d["B2"] = "DESPIECE DE KITS — Dashboard"
ws_d["B2"].font = TITLE_FONT
ws_d.merge_cells("B2:G2")
ws_d["B2"].alignment = Alignment(horizontal="left", vertical="center")
ws_d.row_dimensions[2].height = 26

# Selectores: C4=Kit, C5=Estructura, C6=Marca, C7=Soporte
ws_d["B4"] = "Kit:"
ws_d["B5"] = "Estructura:"
ws_d["B6"] = "Marca Estructura:"
ws_d["B7"] = "Soporte Batería:"
for c in ["B4", "B5", "B6", "B7"]:
    ws_d[c].font = LABEL_FONT
    ws_d[c].alignment = Alignment(horizontal="right", vertical="center")

ws_d["C4"] = "EC-3K"
ws_d["C5"] = "Chapa"
ws_d["C6"] = "SOLARMET"
ws_d["C7"] = "Pared"
for c in ["C4", "C5", "C6", "C7"]:
    ws_d[c].fill = INPUT_FILL
    ws_d[c].font = INPUT_FONT
    ws_d[c].alignment = CENTER
    ws_d[c].border = BORDER

# Info kit
kits_range = f"Kits!$A$2:$F${kits_last_row}"
ws_d["E4"] = "Nombre:"
ws_d["F4"] = f'=IFERROR(VLOOKUP($C$4,{kits_range},2,FALSE),"")'
ws_d["E5"] = "Fase / Tipo:"
ws_d["F5"] = f'=IFERROR(VLOOKUP($C$4,{kits_range},3,FALSE)&" / "&VLOOKUP($C$4,{kits_range},4,FALSE),"")'
ws_d["E6"] = "Potencia / Paneles:"
ws_d["F6"] = f'=IFERROR(VLOOKUP($C$4,{kits_range},5,FALSE)&" kW · "&VLOOKUP($C$4,{kits_range},6,FALSE)&" paneles","")'
for c in ["E4", "E5", "E6"]:
    ws_d[c].font = LABEL_FONT
    ws_d[c].alignment = Alignment(horizontal="right", vertical="center")
for c in ["F4", "F5", "F6"]:
    ws_d[c].font = Font(bold=True)
    ws_d[c].alignment = LEFT
    ws_d[c].border = BORDER
ws_d.merge_cells("F4:G4")
ws_d.merge_cells("F5:G5")
ws_d.merge_cells("F6:G6")

# Validations
dv_kit = DataValidation(type="list", formula1=f"=Kits!$A$2:$A${kits_last_row}", allow_blank=False)
dv_kit.add("C4")
ws_d.add_data_validation(dv_kit)
dv_est = DataValidation(type="list", formula1='"Chapa,Teja,Suelo/Losa"', allow_blank=False)
dv_est.add("C5")
ws_d.add_data_validation(dv_est)
dv_marca = DataValidation(type="list", formula1='"SOLARMET,MIBET"', allow_blank=False)
dv_marca.add("C6")
ws_d.add_data_validation(dv_marca)
dv_sop = DataValidation(type="list", formula1='"Pared,Suelo"', allow_blank=True)
dv_sop.add("C7")
ws_d.add_data_validation(dv_sop)


def section_header(row, text):
    cell = ws_d.cell(row=row, column=2, value=text)
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = HDR_FILL
    ws_d.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def sub_header(row, headers_list, start_col=2):
    for i, h in enumerate(headers_list, start_col):
        cell = ws_d.cell(row=row, column=i, value=h)
        cell.fill = SUB_FILL
        cell.font = SUB_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def style_data_row(row, start_col, end_col, item_col):
    for c in range(start_col, end_col + 1):
        cell = ws_d.cell(row=row, column=c)
        cell.border = BORDER
        cell.alignment = LEFT if c == item_col else CENTER


BOM_BASE_RANGE = f"BOM_Base!$A$2:$H${bom_base_last}"
BOM_EST_RANGE  = f"BOM_Estructura!$A$2:$H${bom_est_last}"
BOM_SOP_RANGE  = f"BOM_Soporte!$A$2:$H${bom_sop_last}"
ED_RANGE       = f"Estructuras_Detalle!$A$2:$G${ed_last}"

# ===== COMPONENTES BASE =====
row = 9
section_header(row, "COMPONENTES BASE DEL KIT")
row += 1
sub_header(row, ["Qty", "Item", "Categoría", "Marca", "Código"])
row += 1
# BOM_Base cols: 4=Qty, 5=Item, 6=Cat, 7=Marca, 8=Cod
for n in range(1, BASE_ROWS_MAX + 1):
    key_expr = f'$C$4&"|"&{n}'
    ws_d.cell(row=row, column=2, value=f'=IFERROR(VLOOKUP({key_expr},{BOM_BASE_RANGE},4,FALSE),"")')
    ws_d.cell(row=row, column=3, value=f'=IFERROR(VLOOKUP({key_expr},{BOM_BASE_RANGE},5,FALSE),"")')
    ws_d.cell(row=row, column=4, value=f'=IFERROR(VLOOKUP({key_expr},{BOM_BASE_RANGE},6,FALSE),"")')
    ws_d.cell(row=row, column=5, value=f'=IFERROR(VLOOKUP({key_expr},{BOM_BASE_RANGE},7,FALSE),"")')
    ws_d.cell(row=row, column=6, value=f'=IFERROR(VLOOKUP({key_expr},{BOM_BASE_RANGE},8,FALSE),"")')
    style_data_row(row, 2, 6, 3)
    row += 1

row += 1

# ===== ESTRUCTURA SELECCIONADA =====
section_header(row, "ESTRUCTURA SELECCIONADA (SET)")
row += 1
sub_header(row, ["Qty", "Item", "Categoría", "Marca", "Código Set"])
row += 1
# Key = C4|C5|C6. BOM_Estructura cols: 5=Qty, 6=Item, 7=Cat, 4=Marca, 8=CodSet
key_est = '$C$4&"|"&$C$5&"|"&$C$6'
ws_d.cell(row=row, column=2, value=f'=IFERROR(VLOOKUP({key_est},{BOM_EST_RANGE},5,FALSE),"")')
ws_d.cell(row=row, column=3, value=f'=IFERROR(VLOOKUP({key_est},{BOM_EST_RANGE},6,FALSE),"")')
ws_d.cell(row=row, column=4, value=f'=IFERROR(VLOOKUP({key_est},{BOM_EST_RANGE},7,FALSE),"")')
ws_d.cell(row=row, column=5, value=f'=IFERROR(VLOOKUP({key_est},{BOM_EST_RANGE},4,FALSE),"")')
ws_d.cell(row=row, column=6, value=f'=IFERROR(VLOOKUP({key_est},{BOM_EST_RANGE},8,FALSE),"")')
style_data_row(row, 2, 6, 3)
set_code_row = row  # F en esta fila = código del set
row += 2

# ===== DESPIECE INTERNO ESTRUCTURA =====
section_header(row, "DESPIECE INTERNO DE LA ESTRUCTURA (por set)")
row += 1
ws_d.cell(row=row, column=2, value="Set:").font = LABEL_FONT
ws_d.cell(row=row, column=2).alignment = Alignment(horizontal="right", vertical="center")
ws_d.cell(row=row, column=3, value=f"=$F${set_code_row}").font = Font(bold=True)
ws_d.cell(row=row, column=3).alignment = CENTER
ws_d.cell(row=row, column=4, value="Sets a usar:").font = LABEL_FONT
ws_d.cell(row=row, column=4).alignment = Alignment(horizontal="right", vertical="center")
ws_d.cell(row=row, column=5, value=f"=$B${set_code_row}").font = Font(bold=True)
ws_d.cell(row=row, column=5).alignment = CENTER
set_ref_row = row
row += 1
sub_header(row, ["Qty x set", "Qty Total", "Item", "Categoría", "Marca", "Código"])
row += 1
# Estructuras_Detalle cols: 4=Qty, 5=Item, 6=Cat, 3=Marca, 7=Cod
for n in range(1, DET_ROWS_MAX + 1):
    key_expr = f'$C${set_ref_row}&"|"&{n}'
    ws_d.cell(row=row, column=2, value=f'=IFERROR(VLOOKUP({key_expr},{ED_RANGE},4,FALSE),"")')
    ws_d.cell(row=row, column=3, value=f'=IFERROR(VLOOKUP({key_expr},{ED_RANGE},4,FALSE)*$E${set_ref_row},"")')
    ws_d.cell(row=row, column=4, value=f'=IFERROR(VLOOKUP({key_expr},{ED_RANGE},5,FALSE),"")')
    ws_d.cell(row=row, column=5, value=f'=IFERROR(VLOOKUP({key_expr},{ED_RANGE},6,FALSE),"")')
    ws_d.cell(row=row, column=6, value=f'=IFERROR(VLOOKUP({key_expr},{ED_RANGE},3,FALSE),"")')
    ws_d.cell(row=row, column=7, value=f'=IFERROR(VLOOKUP({key_expr},{ED_RANGE},7,FALSE),"")')
    style_data_row(row, 2, 7, 4)
    row += 1

# Nota cuando es kit indivisible (MIBET Suelo)
row += 0
nota_row = row
ws_d.cell(row=nota_row, column=2, value=(
    f'=IF($C$6="MIBET",IF($C$5="Suelo/Losa",'
    f'"Kit indivisible: ya viene armado (sin despiece adicional).",""),"")'
))
ws_d.cell(row=nota_row, column=2).font = Font(italic=True, color="7F7F7F")
ws_d.cell(row=nota_row, column=2).alignment = Alignment(horizontal="left", vertical="center")
ws_d.merge_cells(start_row=nota_row, start_column=2, end_row=nota_row, end_column=7)
row += 2

# ===== SOPORTE BATERÍA =====
section_header(row, "SOPORTE BATERÍA (solo kits HV)")
row += 1
sub_header(row, ["Qty", "Item", "Categoría", "Marca", "Código"])
row += 1
key_sop = '$C$4&"|"&$C$7'
ws_d.cell(row=row, column=2, value=f'=IFERROR(VLOOKUP({key_sop},{BOM_SOP_RANGE},4,FALSE),"N/A")')
ws_d.cell(row=row, column=3, value=f'=IFERROR(VLOOKUP({key_sop},{BOM_SOP_RANGE},5,FALSE),"N/A para este kit")')
ws_d.cell(row=row, column=4, value=f'=IFERROR(VLOOKUP({key_sop},{BOM_SOP_RANGE},6,FALSE),"")')
ws_d.cell(row=row, column=5, value=f'=IFERROR(VLOOKUP({key_sop},{BOM_SOP_RANGE},7,FALSE),"")')
ws_d.cell(row=row, column=6, value=f'=IFERROR(VLOOKUP({key_sop},{BOM_SOP_RANGE},8,FALSE),"")')
style_data_row(row, 2, 6, 3)

autosize(ws_d, [2, 12, 48, 14, 14, 14, 22])
ws_d.sheet_view.showGridLines = False

try:
    wb.save(DST)
    print("Saved:", DST)
except PermissionError:
    alt = DST.replace(".xlsx", "_NEW.xlsx")
    wb.save(alt)
    print(f"WARN: {DST} bloqueado (¿abierto en Excel?). Guardado como: {alt}")

print(f"  Kits: {len(KITS)}")
print(f"  BOM_Base rows: {bom_base_last - 1}")
print(f"  BOM_Estructura rows: {bom_est_last - 1} (SOLARMET + MIBET)")
print(f"  BOM_Soporte rows: {bom_sop_last - 1}")
print(f"  Estructuras_Detalle rows: {ed_last - 1}")
